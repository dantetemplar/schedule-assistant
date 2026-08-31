from datetime import time
from pathlib import Path

import yaml
from openpyxl import Workbook, load_workbook

from convert_json_to_config_candidate import (
    PROGRAMS,
    build_sections,
    collapse_spanning_core_rows,
    build_english_entities,
    expand_grouped_core_courses_to_rows,
    extract_elective_entities,
    load_english_schedule,
    merge_schedule_entities,
)
from config import InstructorConfig, SectionConfig
from instructors_roster import InstructorRegistry


def _write_english_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Grouping"
    sheet.append(
        [
            "Доменный идентификатор",
            "Учебная группа",
            "English Course",
            "English Group",
            "Instructor",
            "Instructor email",
            "Day",
            "Time",
            "Room",
        ]
    )
    sheet.append(
        [
            "student@innopolis.university",
            "B26-RO15-01",
            "AWA-I",
            "AWA-I 1",
            "Alina Arslanova",
            "a.arslanova@innopolis.ru",
            "M/W",
            time(12, 40),
            306,
        ]
    )
    workbook.save(path)


def test_load_english_schedule_builds_membership_and_expands_days(
    tmp_path: Path,
) -> None:
    path = tmp_path / "english.xlsx"
    _write_english_workbook(path)

    parsed = load_english_schedule(path)

    assert parsed["diagnostics"] == []
    assert parsed["groups"] == [
        {
            "code": "AWA-I-1",
            "course_name": "AWA-I",
            "students": ["student@innopolis.university"],
            "academic_groups": ["B26-RO-01"],
            "instructor": "Alina Arslanova",
            "instructor_email": "a.arslanova@innopolis.ru",
            "weekly_pattern": [
                {
                    "weekday": "MONDAY",
                    "start_time": "12:40",
                    "end_time": "14:10",
                    "room": "306",
                },
                {
                    "weekday": "WEDNESDAY",
                    "start_time": "12:40",
                    "end_time": "14:10",
                    "room": "306",
                },
            ],
        }
    ]


def test_load_english_schedule_sorts_group_codes_naturally(tmp_path: Path) -> None:
    path = tmp_path / "english.xlsx"
    _write_english_workbook(path)
    workbook = load_workbook(path)
    sheet = workbook["Grouping"]
    for group_code in ("AWA-I 10", "AWA-I 2"):
        sheet.append(
            [
                f"{group_code.lower().replace(' ', '-')}@innopolis.university",
                "B26-RO15-01",
                "AWA-I",
                group_code,
                "Alina Arslanova",
                "a.arslanova@innopolis.ru",
                "M/W",
                time(12, 40),
                306,
            ]
        )
    workbook.save(path)

    parsed = load_english_schedule(path)

    assert [group["code"] for group in parsed["groups"]] == [
        "AWA-I-1",
        "AWA-I-2",
        "AWA-I-10",
    ]


def test_build_english_entities_uses_full_and_short_course_names(
    tmp_path: Path,
) -> None:
    path = tmp_path / "english.xlsx"
    _write_english_workbook(path)
    parsed = load_english_schedule(path)
    instructor = InstructorConfig(
        id="a.arslanova@innopolis.ru",
        name_en="Alina Arslanova",
    )
    instructors = {instructor.id: instructor}
    registry = InstructorRegistry(
        by_id=instructors,
        name_index={"alina arslanova": instructor.id},
    )

    entities = build_english_entities(parsed, instructors, registry)

    assert entities["courses"][0]["name"] == "Academic Writing and Argumentation I"
    assert entities["courses"][0]["short_name"] == "AWA-I"
    assert entities["section"]["programs"][0]["tracks"][0]["name"] == "AWA-I"


def test_english_section_uses_compact_groups_layout() -> None:
    english_section = next(
        section for section in build_sections(PROGRAMS) if section["code"] == "english"
    )

    parsed = SectionConfig.model_validate(english_section)

    assert parsed.default_layout == "compact_groups"


def test_core_rows_exclude_groups_from_the_other_language_program() -> None:
    rows = expand_grouped_core_courses_to_rows(
        [
            {
                "subject": "Course",
                "google_sheet_name": "Ru Programs",
                "components": [
                    {
                        "type": "lab",
                        "weekday": "THURSDAY",
                        "start_time": "12:40",
                        "end_time": "14:10",
                        "audience": ["B24-MFAI-03", "B24-RO-01"],
                    }
                ],
            }
        ]
    )

    assert rows[0]["group_name"] == ["B24-MFAI-03"]


def test_collapse_spanning_core_rows_merges_start_and_end_modifiers() -> None:
    rows = [
        {
            "lesson_name": "Управление данными",
            "lesson_class_type": "лек",
            "weekday": "MONDAY",
            "start_time": "09:00",
            "end_time": "10:30",
            "room": "ОНЛАЙН",
            "teacher": "Armen Beklaryan",
            "group_name": ["B24-MFAI-01", "B24-MFAI-02"],
            "modifiers": {"location": "ОНЛАЙН", "starts_at": "10:00"},
            "a1_range": "S4:T4",
        },
        {
            "lesson_name": "Управление данными",
            "lesson_class_type": "лек",
            "weekday": "MONDAY",
            "start_time": "10:40",
            "end_time": "12:10",
            "room": "ОНЛАЙН",
            "teacher": "Armen Beklaryan",
            "group_name": ["B24-MFAI-01", "B24-MFAI-02"],
            "modifiers": None,
            "a1_range": "S7:T7",
        },
        {
            "lesson_name": "Управление данными",
            "lesson_class_type": "лек",
            "weekday": "MONDAY",
            "start_time": "12:40",
            "end_time": "14:10",
            "room": "ОНЛАЙН",
            "teacher": "Armen Beklaryan",
            "group_name": ["B24-MFAI-01", "B24-MFAI-02"],
            "modifiers": {"location": "ОНЛАЙН", "till": "13:00"},
            "a1_range": "S10:T10",
        },
    ]

    collapsed = collapse_spanning_core_rows(rows)

    assert len(collapsed) == 1
    assert collapsed[0]["start_time"] == "10:00"
    assert collapsed[0]["end_time"] == "13:00"
    assert collapsed[0]["modifiers"] == {"location": "ОНЛАЙН"}
    assert collapsed[0]["a1_range"] == "S4:T4;S7:T7;S10:T10"


def test_extract_electives_keeps_only_linked_entities(tmp_path: Path) -> None:
    path = tmp_path / "mixed.yaml"
    path.write_text(
        yaml.safe_dump(
            {
                "term": {
                    "sections": [
                        {"code": "core", "name": "Core"},
                        {"code": "electives", "name": "Electives"},
                    ]
                },
                "courses": [
                    {"name": "Core", "section_code": "core", "components": []},
                    {
                        "name": "Elective",
                        "section_code": "electives",
                        "components": [
                            {
                                "tag": "class",
                                "audience": ["EL"],
                                "instructor_pool": ["elective@example.com"],
                                "sessions": [
                                    {
                                        "weekly_pattern": [
                                            {
                                                "room": "E1",
                                                "instructor": "elective@example.com",
                                            }
                                        ]
                                    }
                                ],
                            }
                        ],
                    },
                ],
                "students_groups": [
                    {"code": "CORE", "kind": "core"},
                    {"code": "EL", "kind": "elective"},
                ],
                "rooms": [{"id": "C1"}, {"id": "E1"}],
                "instructors": [
                    {"id": "core@example.com", "slot_preferences": []},
                    {"id": "elective@example.com", "slot_preferences": []},
                ],
            },
            sort_keys=False,
        ),
        encoding="utf-8",
    )

    extracted = extract_elective_entities(path)

    assert [item["code"] for item in extracted["sections"]] == ["electives"]
    assert [item["name"] for item in extracted["courses"]] == ["Elective"]
    assert [item["code"] for item in extracted["students_groups"]] == ["EL"]
    assert [item["id"] for item in extracted["rooms"]] == ["E1"]
    assert extracted["instructors"] == [{"id": "elective@example.com"}]


def test_keyed_merge_replaces_source_owned_entities_idempotently() -> None:
    generated = {
        "term": {"sections": [{"code": "core"}, {"code": "english"}]},
        "students_groups": [
            {"code": "CORE", "kind": "core"},
            {"code": "ENG", "kind": "english", "students": []},
        ],
        "courses": [
            {"name": "Core", "section_code": "core", "components": []},
            {"name": "English", "section_code": "english", "components": []},
        ],
        "rooms": [{"id": "101"}],
        "instructors": [{"id": "core@example.com"}],
    }
    english = {
        "section": {"code": "english", "programs": []},
        "students_groups": [
            {"code": "ENG", "kind": "english", "students": ["student"]}
        ],
        "courses": [
            {
                "name": "English",
                "section_code": "english",
                "components": [{"tag": "class"}],
            }
        ],
        "instructors": [{"id": "english@example.com"}],
    }
    electives = {
        "sections": [{"code": "electives"}],
        "students_groups": [{"code": "EL", "kind": "elective"}],
        "courses": [
            {"name": "Elective", "section_code": "electives", "components": []}
        ],
        "rooms": [{"id": "E1"}],
        "instructors": [
            {
                "id": "elective@example.com",
                "position": "Docent",
                "slot_preferences": [],
            },
            {"id": "ta@example.com", "position": "TA"},
            {"id": "professor@example.com", "position": "Professor"},
        ],
    }

    first = merge_schedule_entities(generated, english, electives)
    second = merge_schedule_entities(first, english, electives)

    assert first == second
    assert len({group["code"] for group in first["students_groups"]}) == 3
    assert next(
        group for group in first["students_groups"] if group["code"] == "ENG"
    )["students"] == ["student"]
    assert {
        instructor["id"]: instructor.get("position")
        for instructor in first["instructors"]
    } == {
        "core@example.com": None,
        "english@example.com": None,
        "elective@example.com": "Associate Professor",
        "ta@example.com": "Teaching Assistant",
        "professor@example.com": "Full Professor",
    }
    assert "slot_preferences" not in str(first)
