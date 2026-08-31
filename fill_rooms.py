#!/usr/bin/env -S uv run --script
# /// script
# requires-python = ">=3.14"
# dependencies = [
#     "openpyxl>=3.1.5",
# ]
# ///

input_filename = "BS intake 2026 grouping v2 290826.xlsx"
output_filename = "English Groups.xlsx"

from datetime import datetime, time

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell


instructor_names = {
    "Valeria": "Valeria Tishkova",
    "Alexandra": "Alexandra Vasilieva",
    "Alina": "Alina Arslanova",
    "Anna": "Anna Startseva",
    "Elvira": "Elvira Kharassova",
    "Evgenia": "Evgenia Kruglova",
    "Georgy": "Georgy Gelvanovsky",
    "Irina": "Irina Rednikova",
    "Kamila": "Kamilla Sakhabieva",
    "Kamilla": "Kamilla Sakhabieva",
    "Maria": "Maria Melnikova",
    "Ruslan": "Ruslan Saduov",
}

instructor_emails = {
    "Valeria Tishkova": "v.tishkova@innopolis.ru",
    "Alexandra Vasilieva": "a.vasilyeva@innopolis.ru",
    "Alina Arslanova": "a.arslanova@innopolis.ru",
    "Anna Startseva": "an.startseva@innopolis.university",
    "Elvira Kharassova": "e.kharrasova@innopolis.ru",
    "Evgenia Kruglova": "e.kruglova@innopolis.ru",
    "Georgy Gelvanovsky": "g.gelvanovsky@innopolis.ru",
    "Irina Rednikova": "i.rednikova@innopolis.ru",
    "Kamilla Sakhabieva": "k.sakhabieva@innopolis.ru",
    "Maria Melnikova": "m.melnikova@innopolis.ru",
    "Ruslan Saduov": "ru.saduov@innopolis.ru",
}

rooms = {
    "M/W 12:40 Alexandra": 102,
    "M/W 14:20 Alexandra": 102,
    "M/W 16:00 Alexandra": 102,
    "M/W 12:40 Alina": 306,
    "M/W 14:20 Alina": 306,
    "M/W 16:00 Alina": 306,
    "T/Th 14:20 Alina": 306,
    "M/W 12:40 Anna": 304,
    "M/W 14:20 Anna": 304,
    "T/Th 14:20 Anna": 102,
    "T/Th 16:00 Anna": 102,
    "M/W 12:40 Evgenia": 104,
    "M/W 14:20 Evgenia": 104,
    "T/Th 14:20 Evgenia": 104,
    "T/Th 16:00 Evgenia": 104,
    "M/W 14:20 Elvira": 305,
    "M/W 16:00 Elvira": 305,
    "T/Th 14:20 Elvira": 305,
    "T/Th 16:00 Elvira": 305,
    "M/W 12:40 Georgy": 111,
    "M/W 14:20 Georgy": 111,
    "M/W 16:00 Georgy": 111,
    "M/W 12:40 Irina": 301,
    "M/W 14:20 Irina": 301,
    "T/Th 14:20 Irina": 301,
    "T/Th 16:00 Irina": 301,
    "T/Th 14:20 Kamila": 103,
    "T/Th 16:00 Kamila": 103,
    "M/W 14:20 Maria": 103,
    "M/W 16:00 Maria": 103,
    "T/Th 14:20 Ruslan": 304,
    "T/Th 16:00 Ruslan": 304,
}

room_instructor_names = {
    full_name: short_name
    for short_name, full_name in instructor_names.items()
    if short_name != "Kamilla"
    and any(key.endswith(f" {short_name}") for key in rooms)
}


def format_time(value: object) -> str:
    if isinstance(value, (datetime, time)):
        return value.strftime("%H:%M")
    return str(value).strip()


def main() -> None:
    workbook = load_workbook(input_filename)
    filled_rooms = 0

    for sheet in workbook.worksheets:
        header_row = None
        columns = {}

        # The people list ends at Room; columns to the right are unrelated.
        for row in sheet.iter_rows(max_col=16):
            values = {
                str(cell.value).strip(): cell.column
                for cell in row
                if cell.value is not None
            }
            required_columns = {"Instructor", "Day", "Time", "Room"}
            if required_columns <= values.keys():
                header_row = row[0].row
                columns = values
                break

        if header_row is None:
            continue

        if "Instructor email" not in columns:
            instructor_column = columns["Instructor"]
            sheet.insert_cols(instructor_column + 1)
            email_header_cell = sheet.cell(header_row, instructor_column + 1)
            if not isinstance(email_header_cell, Cell):
                raise TypeError("Unexpected merged Instructor email header cell")
            email_header_cell.value = "Instructor email"
            columns = {
                str(cell.value).strip(): cell.column
                for cell in sheet[header_row]
                if cell.value is not None
            }

        for row_number in range(header_row + 1, sheet.max_row + 1):
            instructor = sheet.cell(row_number, columns["Instructor"]).value
            day = sheet.cell(row_number, columns["Day"]).value
            lesson_time = sheet.cell(row_number, columns["Time"]).value

            if not instructor and not day and not lesson_time:
                continue
            if not instructor or not day or lesson_time is None:
                continue

            instructor_token = str(instructor).strip()
            room_instructor = room_instructor_names.get(
                instructor_token, instructor_token
            )
            key = (
                f"{str(day).strip()} {format_time(lesson_time)} "
                f"{room_instructor}"
            )
            if key not in rooms:
                raise KeyError(f"Room is not specified for row {row_number}: {key}")

            room_cell = sheet.cell(row_number, columns["Room"])
            instructor_cell = sheet.cell(row_number, columns["Instructor"])
            if not isinstance(room_cell, Cell) or not isinstance(instructor_cell, Cell):
                raise TypeError(f"Unexpected merged cell in row {row_number}")
            room_cell.value = rooms[key]
            full_name = instructor_names.get(
                instructor_token,
                instructor_token if instructor_token in room_instructor_names else None,
            )
            if full_name is None:
                raise KeyError(
                    f"Full instructor name is not specified for row {row_number}: "
                    f"{instructor}"
                )
            instructor_email = instructor_emails.get(full_name)
            if instructor_email is None:
                raise KeyError(
                    f"Instructor email is not specified for row {row_number}: "
                    f"{full_name}"
                )
            instructor_cell.value = full_name
            email_cell = sheet.cell(row_number, columns["Instructor email"])
            if not isinstance(email_cell, Cell):
                raise TypeError(f"Unexpected merged cell in row {row_number}")
            email_cell.value = instructor_email
            filled_rooms += 1

    if filled_rooms == 0:
        raise RuntimeError("No rows from the people list were found")

    workbook.save(output_filename)
    print(f"Filled {filled_rooms} rooms. Saved to {output_filename}")


if __name__ == "__main__":
    main()